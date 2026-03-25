from flask import Flask, render_template, request, jsonify, send_file, send_from_directory, make_response
import os
import io
import re
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook, load_workbook
import dateutil.parser

# Optional GCS support
GCS_BUCKET = os.getenv("GCS_BUCKET", "").strip()
GCS_ENABLED = bool(GCS_BUCKET)
if GCS_ENABLED:
    from google.cloud import storage

app = Flask(__name__)
app.config['SECRET_KEY'] = 'marmoset'


class DataLogger:
    def __init__(self):
        # Local storage (fallback when GCS not enabled)
        self.config_dir = os.path.join(os.path.expanduser('~'), 'DataLogApp')
        os.makedirs(self.config_dir, exist_ok=True)
        self.counter_file = os.path.join(self.config_dir, 'sample_name_counter.json')

        # GCS client
        self.storage_client = storage.Client() if GCS_ENABLED else None

        self.name_to_code = {
            "Petra": "CJ23.56.001",
            "Croissant": "CJ23.56.002",
            "Nutmeg": "CJ23.56.003",
            "Tank": "CJ23.56.004",
            "JellyBean": "CJ24.56.001",
            "Pringle": "CJ24.56.002",
            "Paarl": "CJ24.56.003",
            "Rambo": "CJ24.56.004",
            "Clack": "CJ24.56.005",
            "Porthos": "CJ24.56.006",
            "Deegan": "CJ24.56.007",
            "Dangerboy": "CJ24.56.008",
            "Hildegard": "CJ24.56.009",
            "Villopoto": "CJ24.56.010",
            "Pathy": "CJ24.56.011",
            "Toki": "CJ24.56.012",
            "Georgia": "CJ24.56.013",
            "Carmichael": "CJ24.56.014",
            "Morel": "CJ24.56.015",
            "Orion": "CJ24.56.016",
            "EllieMae": "CJ24.56.017",
            "Lambert": "CJ24.56.018",
            "Ocean": "CJ25.56.001",
            "Stella": "CJ25.56.002",
            "Wyatt": "CJ25.56.003",
            "Piglet": "CJ25.56.004",
            "Moira": "CJ25.56.005",
            "Willow": "CJ25.56.006",
            "Wren": "CJ25.56.007",
            "Valentino": "CJ25.56.008",
            "Misty": "CJ25.56.009",
            "Link": "CJ25.56.010",
            "Owlette": "CJ25.56.011",
            "Chickpea": "CJ25.56.012",
            "Benedict": "CJ25.56.013",
            "Vera": "CJ25.56.014",
            "Tango": "CJ25.56.015",
            "Paris": "CJ25.56.016",
            "Lapras": "CJ25.56.017"
        }

        self.black_fill = PatternFill(start_color='000000', fill_type='solid')

    # ----------------- User key and object names -----------------

    def _safe_user_key(self, name: str) -> str:
        name = (name or "").strip()
        if not name:
            return "unknown"
        return re.sub(r'[^A-Za-z0-9_-]+', '_', name) or "unknown"

    # ----------------- Workbook storage helpers (GCS/local) -----------------

    def _new_object_name(self, user_key: str) -> str:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"logs/{user_key}/{user_key}_krienen_data_log_{ts}.xlsx"

    def _load_pointer(self, user_key: str):
        if GCS_ENABLED:
            bucket = self.storage_client.bucket(GCS_BUCKET)
            blob = bucket.blob(f"pointers/{user_key}.json")
            if blob.exists():
                try:
                    data = json.loads(blob.download_as_text())
                    if isinstance(data, dict) and data.get("object"):
                        return data["object"]
                except Exception:
                    pass
        # local fallback mapping (optional)
        mapping = self._load_local_meta().get("current_log_objects", {})
        return mapping.get(user_key)

    def _save_pointer(self, user_key: str, object_name: str):
        if GCS_ENABLED:
            bucket = self.storage_client.bucket(GCS_BUCKET)
            blob = bucket.blob(f"pointers/{user_key}.json")
            blob.upload_from_string(json.dumps({"object": object_name}, indent=2), content_type="application/json")
        meta = self._load_local_meta()
        meta.setdefault("current_log_objects", {})[user_key] = object_name
        self._save_local_meta(meta)

    def _download_workbook(self, object_name):
        if GCS_ENABLED:
            bucket = self.storage_client.bucket(GCS_BUCKET)
            blob = bucket.blob(object_name)
            if blob.exists():
                data = blob.download_as_bytes()
                wb = load_workbook(io.BytesIO(data))
                return wb, blob.generation
            else:
                wb = self._initialize_excel()
                return wb, None
        else:
            local_path = os.path.join(self.config_dir, object_name.replace("/", os.sep))
            os.makedirs(os.path.dirname(local_path), exist_ok=True)
            if os.path.exists(local_path):
                wb = load_workbook(local_path)
            else:
                wb = self._initialize_excel()
            return wb, None

    def _upload_workbook(self, wb, object_name, if_generation_match=None):
        if GCS_ENABLED:
            bucket = self.storage_client.bucket(GCS_BUCKET)
            blob = bucket.blob(object_name)
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            if if_generation_match is not None:
                blob.upload_from_file(
                    out,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    if_generation_match=if_generation_match
                )
            else:
                blob.upload_from_file(
                    out,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
        else:
            local_path = os.path.join(self.config_dir, object_name.replace("/", os.sep))
            os.makedirs(os.path.dirname(local_path), exist_ok=True)
            wb.save(local_path)

    def _download_workbook_bytes(self, object_name):
        if GCS_ENABLED:
            bucket = self.storage_client.bucket(GCS_BUCKET)
            blob = bucket.blob(object_name)
            if blob.exists():
                return blob.download_as_bytes()
            else:
                wb = self._initialize_excel()
                out = io.BytesIO()
                wb.save(out)
                return out.getvalue()
        else:
            local_path = os.path.join(self.config_dir, object_name.replace("/", os.sep))
            if os.path.exists(local_path):
                with open(local_path, 'rb') as f:
                    return f.read()
            else:
                wb = self._initialize_excel()
                out = io.BytesIO()
                wb.save(out)
                return out.getvalue()

    # ----------------- Per-user state (local meta fallback only) -----------------

    def _load_local_meta(self):
        if os.path.exists(self.counter_file):
            try:
                with open(self.counter_file, 'r') as f:
                    return json.load(f)
            except Exception:
                return {}
        return {}

    def _save_local_meta(self, meta: dict):
        with open(self.counter_file, 'w') as f:
            json.dump(meta, f, indent=4)

    # ----------------- Core utilities -----------------

    def _headers(self):
        return ['krienen_lab_identifier', 'seq_portal', 'elab_link', 'experiment_start_date',
                'mit_name', 'donor_name', 'tissue_name', 'tissue_name_old',
                'dissociated_cell_sample_name', 'facs_population_plan', 'cell_prep_type',
                'study', 'enriched_cell_sample_container_name', 'expc_cell_capture',
                'port_well', 'enriched_cell_sample_name', 'enriched_cell_sample_quantity_count',
                'barcoded_cell_sample_name', 'library_method', 'cDNA_amplification_method',
                'cDNA_amplification_date', 'amplified_cdna_name', 'cDNA_pcr_cycles',
                'rna_amplification_pass_fail', 'percent_cdna_longer_than_400bp',
                'cdna_amplified_quantity_ng', 'cDNA_library_input_ng', 'library_creation_date',
                'library_prep_set', 'library_name', 'tapestation_avg_size_bp',
                'library_num_cycles', 'lib_quantification_ng', 'library_prep_pass_fail',
                'r1_index', 'r2_index', 'ATAC_index']

    def _initialize_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "HMBA"
        ws.append(self._headers())
        for col_num in range(1, len(self._headers()) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(name="Arial", size=10, bold=True)
            cell.alignment = Alignment(horizontal='left')
        return wb

    def convert_date(self, exp_date):
        clean = "".join(c for c in exp_date if c.isdigit())
        if len(clean) == 6:
            try:
                datetime.strptime(clean, '%y%m%d')
                return clean
            except ValueError:
                pass
        try:
            parsed = dateutil.parser.parse(exp_date)
            return parsed.strftime('%y%m%d')
        except ValueError:
            return None

    def convert_index(self, index):
        index = index.strip().upper()
        if len(index) == 3:
            if index[0].isdigit() and index[1].isdigit() and index[2].isalpha():
                return f"{index[2]}{index[0]}{index[1]}"
            elif index[0].isalpha() and index[1].isdigit() and index[2].isdigit():
                return index
        elif len(index) == 2:
            if index[0].isdigit() and index[1].isalpha():
                return f"{index[1]}0{index[0]}"
            elif index[0].isalpha() and index[1].isdigit():
                return f"{index[0]}0{index[1]}"
        return None

    def pad_index(self, index):
        if len(index) == 2 and index[0].isalpha() and index[1].isdigit():
            return f"{index[0]}0{index[1]}"
        return index

    # ----------------- Sheet-derived state helpers -----------------

    def _sheet_max_chip(self, ws):
        """
        Scan the entire sheet to find the highest chip number (P####) across
        ALL dates.  Returns 0 if no barcoded_cell_sample_name rows exist.
        """
        headers = [cell.value for cell in ws[1]]
        bcsn_col = headers.index('barcoded_cell_sample_name') + 1
        max_chip = 0
        for row_idx in range(2, ws.max_row + 1):
            name_val = ws.cell(row=row_idx, column=bcsn_col).value
            if not name_val or not isinstance(name_val, str):
                continue
            m = re.match(r'^P(\d{4})_(\d+)$', name_val)
            if m:
                max_chip = max(max_chip, int(m.group(1)))
        return max_chip

    def _sheet_date_chip_usage(self, ws, current_date):
        """
        Scan the sheet for rows with experiment_start_date == current_date
        and build a map of chip -> highest used well for that date.
        barcoded_cell_sample_name format: 'P####_##'
        """
        headers = [cell.value for cell in ws[1]]
        date_col = headers.index('experiment_start_date') + 1
        bcsn_col = headers.index('barcoded_cell_sample_name') + 1

        chips_map = {}  # chip_str -> used_wells (int)
        last_chip = None
        last_used = 0

        for row_idx in range(2, ws.max_row + 1):
            date_val = ws.cell(row=row_idx, column=date_col).value
            if date_val != current_date:
                continue
            name_val = ws.cell(row=row_idx, column=bcsn_col).value
            if not name_val or not isinstance(name_val, str):
                continue
            m = re.match(r'^P(\d{4})_(\d+)$', name_val)
            if not m:
                continue
            chip = int(m.group(1))
            well = int(m.group(2))
            chips_map[str(chip)] = max(well, int(chips_map.get(str(chip), 0)))
        if chips_map:
            # Highest chip that has entries on this date
            last_chip = max(int(c) for c in chips_map.keys())
            last_used = int(chips_map[str(last_chip)])
        return chips_map, last_chip, last_used

    def _next_amp_name(self, ws, amp_prefix, amp_date):
        """
        Determine the next amplified_cdna_name by scanning existing rows in the sheet:
        pattern: f"{amp_prefix}_{amp_date}_{batch}_{letter}"
        letter cycles A..H; after H, batch increments.
        """
        headers = [cell.value for cell in ws[1]]
        amp_name_col = headers.index('amplified_cdna_name') + 1

        last_batch = 0
        last_letter = None  # 'A'..'H'

        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=amp_name_col).value
            if not val or not isinstance(val, str):
                continue
            # Example: APLCTX_251001_1_G
            m = re.match(rf'^{re.escape(amp_prefix)}_{re.escape(amp_date)}_(\d+)_([A-H])$', val)
            if not m:
                continue
            b = int(m.group(1))
            L = m.group(2)
            if b > last_batch or (b == last_batch and (last_letter is None or L > last_letter)):
                last_batch = b
                last_letter = L

        if last_batch == 0:
            # No prior reactions for this amp date
            return f"{amp_prefix}_{amp_date}_1_A"

        # Compute next letter/batch
        letters = 'ABCDEFGH'
        if last_letter is None:
            # Shouldn't happen but default safely
            return f"{amp_prefix}_{amp_date}_{last_batch}_A"
        idx = letters.index(last_letter)
        if idx < 7:
            next_letter = letters[idx + 1]
            next_batch = last_batch
        else:
            next_letter = 'A'
            next_batch = last_batch + 1

        return f"{amp_prefix}_{amp_date}_{next_batch}_{next_letter}"

    # ----------------- Business logic -----------------

    # ----------------- Business logic -----------------

    def process_form_data(self, form_data):
        # Import heavy modules only when needed
        from openpyxl.utils import get_column_letter

        # 1. Setup user workbook and state
        user_first_name = form_data.get('user_first_name', '').strip()
        user_key = self._safe_user_key(user_first_name)

        object_name = self._load_pointer(user_key)
        if not object_name:
            object_name = self._new_object_name(user_key)
            self._save_pointer(user_key, object_name)

        workbook, generation = self._download_workbook(object_name)
        worksheet = workbook.active

        # Determine actual last row
        last_row_with_content = 1
        for row_idx in range(1, worksheet.max_row + 1):
            for cell in worksheet[row_idx]:
                if cell.value is not None:
                    last_row_with_content = row_idx
                    break
        current_row = last_row_with_content + 1

        # Load web app user state
        meta = self._load_local_meta()
        states = meta.setdefault('user_states', {})
        state = states.setdefault(user_key, {"next_counter": 90, "date_info": {}, "amp_counter": {}})
        if "date_info" not in state: state["date_info"] = {}
        if "amp_counter" not in state: state["amp_counter"] = {}
        if "next_counter" not in state or state["next_counter"] is None: state["next_counter"] = 90

        # Get values from form_data dictionary
        current_date = self.convert_date(form_data.get('date', ''))
        mit_name_input = form_data.get('marmoset', '')
        mit_name = "cj" + mit_name_input
        donor_name = self.name_to_code.get(mit_name_input, mit_name_input)

        slab_raw = form_data.get('slab', '').strip()
        hemisphere = form_data.get('hemisphere', '').split()[0].upper() if form_data.get('hemisphere') else ''

        # Split comma-separated slabs and process each one
        slab_parts = [s.strip() for s in slab_raw.split(',') if s.strip()]
        processed_slabs = []
        for s in slab_parts:
            if hemisphere == "RIGHT":
                try:
                    s = str(int(s) + 40).zfill(2)
                except:
                    pass
            elif hemisphere == "BOTH":
                try:
                    s = str(int(s) + 90).zfill(2)
                except:
                    pass
            else:
                if s.isdigit(): s = s.zfill(2)
            processed_slabs.append(s)

        # For tissue_name: join multiple slabs with underscores (e.g. 07_08_09_10)
        slab_for_tissue = "_".join(processed_slabs)
        # For krienen_lab_identifier: keep using the first slab only
        slab = processed_slabs[0] if processed_slabs else slab_raw

        tile_value = form_data.get('tile', '').strip()
        tile = str(int(tile_value)).zfill(2) if tile_value.isdigit() else tile_value

        tile_location_abbr = form_data.get('tile_location', '')

        sort_method = form_data.get('sort_method', '')
        sort_method = sort_method.upper() if sort_method.lower() == "dapi" else sort_method

        if sort_method.lower() == "pooled":
            facs_population = form_data.get('facs_population', '')
        elif sort_method.lower() == "unsorted":
            facs_population = "no_FACS"
        else:
            facs_population = "DAPI"

        try:
            rxn_number = int(form_data.get('rxn_number', 1))
        except ValueError:
            rxn_number = 1

        # ==========================================
        # UPDATED LOGIC FOR COLUMN R (PXXXX counter)
        # Uses sheet-scanning as source of truth so
        # counters survive server restarts / state loss.
        # ==========================================

        # --- Reconcile global next_counter with sheet if state looks default ---
        if state.get("next_counter", 90) == 90 and not state.get("date_info"):
            global_max_chip = self._sheet_max_chip(worksheet)
            if global_max_chip >= 90:
                state["next_counter"] = global_max_chip + 1

        # --- Reconcile state with what's actually in the sheet ---
        chips_map, sheet_last_chip, sheet_last_used = self._sheet_date_chip_usage(worksheet, current_date)

        # Compute total reactions from sheet: sum of max wells across all chips for this date
        sheet_total_reactions = sum(int(v) for v in chips_map.values()) if chips_map else 0
        # The base chip is the lowest chip number used on this date
        sheet_base_chip = min(int(c) for c in chips_map.keys()) if chips_map else None

        if current_date not in state["date_info"]:
            if sheet_base_chip is not None:
                # Sheet already has data for this date that state doesn't know about.
                state["date_info"][current_date] = {
                    "p_number": sheet_base_chip,  # base chip for this date
                    "total_reactions": sheet_total_reactions
                }
                # Ensure global counter stays ahead of any chip number in the sheet
                state["next_counter"] = max(state.get("next_counter", 90), sheet_last_chip + 1)
            else:
                # Truly new date — no data in sheet or state
                p_number = state.get("next_counter", 90)
                state["date_info"][current_date] = {
                    "p_number": p_number,  # base chip for this date
                    "total_reactions": 0
                }
                state["next_counter"] = p_number + 1
        else:
            # State exists for this date — validate against sheet in case of drift
            date_entry = state["date_info"][current_date]
            if sheet_base_chip is not None:
                if sheet_total_reactions > date_entry["total_reactions"]:
                    date_entry["total_reactions"] = sheet_total_reactions
                # Keep base chip aligned with sheet
                if sheet_base_chip != date_entry.get("p_number"):
                    date_entry["p_number"] = sheet_base_chip
                state["next_counter"] = max(state.get("next_counter", 90), sheet_last_chip + 1)

        date_entry = state["date_info"][current_date]

        # OVERRIDE CHECK: Did the user manually update the counter from the UI?
        expected_next = date_entry["p_number"] + 1
        current_global_counter = state.get("next_counter", 90)

        # If they don't match, it means you manually clicked the button to change it
        if current_global_counter != expected_next:
            date_entry["p_number"] = current_global_counter
            # Increment the global counter so the next new date knows what to do
            state["next_counter"] = current_global_counter + 1

        existing_total = date_entry["total_reactions"]
        base_p_number = date_entry["p_number"]

        # Port_well cycles 1-8; after 8, chip number increments
        port_wells = []
        for x in range(rxn_number):
            absolute_idx = existing_total + x  # 0-indexed reaction count
            p_number = base_p_number + (absolute_idx // 8)
            port_well = (absolute_idx % 8) + 1  # 1-8
            port_wells.append((p_number, port_well))

        date_entry["total_reactions"] = existing_total + rxn_number
        # Ensure next_counter stays ahead of the highest chip we just used
        highest_chip_used = base_p_number + ((existing_total + rxn_number - 1) // 8)
        state["next_counter"] = max(state.get("next_counter", 90), highest_chip_used + 1)
        # ==========================================

        atac_indices_raw = form_data.get('atac_indices', '')
        atac_indices = [self.pad_index(self.convert_index(i)) if self.convert_index(i) else i for i in
                        atac_indices_raw.split(',')] if atac_indices_raw else []

        rna_indices_raw = form_data.get('rna_indices', '')
        rna_indices = [self.pad_index(self.convert_index(i)) if self.convert_index(i) else i for i in
                       rna_indices_raw.split(',')] if rna_indices_raw else []

        seq_portal = "no"
        elab_link = form_data.get('elab_link', '')
        study = form_data.get('project', 'HMBA_CjAtlas_Subcortex')
        is_aim4 = (study == 'HMBA_Aim4')

        tissue_name = f"{donor_name}.{tile_location_abbr}.{slab_for_tissue}.{tile}"
        sample_suffix = "Rseq" if is_aim4 else "Multiome"
        dissociated_cell_sample_name = f'{current_date}_{tissue_name}.{sample_suffix}'
        cell_prep_type = "nuclei"

        sorting_status = "PS" if sort_method.lower() in ["pooled", "dapi"] else "PN"
        sorter_initials = form_data.get('sorter_initials', '').strip().upper()
        enriched_prefix = "MPTX" if is_aim4 else "MPXM"
        enriched_cell_sample_container_name = f"{enriched_prefix}_{current_date}_{sorting_status}_{sorter_initials}"

        # Aim4 = RNA only; Multiome = RNA + ATAC
        modalities = ["RNA"] if is_aim4 else ["RNA", "ATAC"]

        dup_index_counter = {}
        headers = [cell.value for cell in worksheet[1]]

        for x in range(rxn_number):
            p_number, port_well = port_wells[x]
            barcoded_cell_sample_name = f'P{str(p_number).zfill(4)}_{port_well}'

            for modality in modalities:
                self.write_modality_data(
                    worksheet, current_row, modality, x, current_date, mit_name, slab_for_tissue, tile, sort_method,
                    port_well, barcoded_cell_sample_name, form_data, tissue_name, rna_indices,
                    atac_indices, headers, dup_index_counter, donor_name, study, state
                )
                current_row += 1

        self._upload_workbook(workbook, object_name, generation if GCS_ENABLED else None)
        self._save_local_meta(meta)
        return True

    def write_modality_data(self, worksheet, current_row, modality, x, current_date, mit_name, slab, tile, sort_method,
                            port_well, barcoded_cell_sample_name, form_data, tissue_name_base, rna_indices,
                            atac_indices, headers, dup_index_counter, donor_name, project, state):

        # slab may be underscore-joined for multi-slab (e.g. "07_08_09_10")
        slab_part = f"Slab{slab}"
        tile_part = f"Tile{int(tile)}" if str(tile).isdigit() else tile

        krienen_lab_identifier = f"{current_date}_HMBA_{mit_name}_{slab_part}_{tile_part}_{sort_method}_{modality}{x + 1}"

        experimenter_initials = form_data.get('sorter_initials', '').strip().upper()
        sorting_status = "PS" if sort_method.lower() in ["pooled", "dapi"] else "PN"
        tissue_name = tissue_name_base

        is_aim4 = (project == 'HMBA_Aim4')
        sample_suffix = "Rseq" if is_aim4 else "Multiome"
        dissociated_cell_sample_name = f'{current_date}_{tissue_name}.{sample_suffix}'
        enriched_prefix = "MPTX" if is_aim4 else "MPXM"
        rna_suffix = "TX" if is_aim4 else "XR"
        atac_suffix = "XA"

        enriched_cell_sample_container_name = f"{enriched_prefix}_{current_date}_{sorting_status}_{experimenter_initials}"
        enriched_cell_sample_name = f'{enriched_prefix}_{current_date}_{sorting_status}_{experimenter_initials}_{port_well}'

        study = project
        seq_portal = "no"
        elab_link = form_data.get('elab_link', '')
        facs_population = form_data.get('facs_population', 'no_FACS')
        cell_prep_type = "nuclei"

        library_prep_date = (self.convert_date(form_data.get('rna_prep_date', '')) if modality == "RNA"
                             else self.convert_date(form_data.get('atac_prep_date', '')))

        # Helper methods to safely extract comma separated inputs from the Web UI
        def safe_float_split(val, idx):
            try:
                return float(str(val).split(',')[idx].strip())
            except:
                return 0.0

        def safe_int_split(val, idx):
            try:
                return int(str(val).split(',')[idx].strip())
            except:
                return 0

        if modality == "RNA":
            library_method = "10xV4" if is_aim4 else "10xMultiome-RSeq"
            library_type = f"LP{experimenter_initials}{rna_suffix}"
            library_index = rna_indices[x] if x < len(rna_indices) else ""

            cdna_concentration = safe_float_split(form_data.get('cdna_concentration', ''), x)
            cdna_amplified_quantity = cdna_concentration * 40
            cdna_library_input = cdna_amplified_quantity * 0.25
            percent_cdna_400bp = safe_float_split(form_data.get('percent_cdna_400bp', ''), x)
            rna_concentration = safe_float_split(form_data.get('rna_lib_concentration', ''), x)
            lib_quant = rna_concentration * 35

            cdna_pcr_cycles = safe_int_split(form_data.get('cdna_pcr_cycles', ''), x)
            rna_size = safe_int_split(form_data.get('rna_sizes', ''), x)
            library_cycles = safe_int_split(form_data.get('library_cycles_rna', ''), x)
        else:
            library_method = "10xMultiome-ASeq"  # ATAC only for Multiome, not Aim4
            library_type = f"LP{experimenter_initials}{atac_suffix}"
            library_index = atac_indices[x] if x < len(atac_indices) else ""

            atac_concentration = safe_float_split(form_data.get('atac_lib_concentration', ''), x)
            lib_quant = atac_concentration * 20

            atac_size = safe_int_split(form_data.get('atac_sizes', ''), x)
            library_cycles = safe_int_split(form_data.get('library_cycles_atac', ''), x)

            cdna_concentration = None
            cdna_amplified_quantity = None
            cdna_library_input = None
            percent_cdna_400bp = None
            cdna_pcr_cycles = None
            rna_size = None

        key = (library_type, library_prep_date, library_index)
        dup_index_counter[key] = dup_index_counter.get(key, 0) + 1
        library_prep_set = f"{library_type}_{library_prep_date}_{dup_index_counter[key]}"
        library_name = f"{library_prep_set}_{library_index}"

        try:
            expected_cell_capture = int(form_data.get('expected_recovery', 0))
        except ValueError:
            expected_cell_capture = 0

        try:
            concentration = float(form_data.get('nuclei_concentration', '0').replace(",", ""))
            volume = float(form_data.get('nuclei_volume', '0'))
            enriched_cell_sample_quantity_count = round(concentration * volume)
        except ValueError:
            enriched_cell_sample_quantity_count = 0

        row_data = [
            krienen_lab_identifier,
            seq_portal,
            elab_link,
            current_date,
            mit_name,
            donor_name,
            tissue_name,
            None,
            dissociated_cell_sample_name,
            facs_population,
            cell_prep_type,
            study,
            enriched_cell_sample_container_name,
            expected_cell_capture,
            port_well,
            enriched_cell_sample_name,
            enriched_cell_sample_quantity_count,
            barcoded_cell_sample_name,
            library_method,
            ("10xV4" if is_aim4 else "10xMultiome-RSeq") if modality == "RNA" else None,
            self.convert_date(form_data.get('cdna_amp_date', '')) if modality == "RNA" else None,
            None,
            cdna_pcr_cycles if modality == "RNA" else None,
            "Pass" if modality == "RNA" else None,
            percent_cdna_400bp if modality == "RNA" else None,
            cdna_amplified_quantity if modality == "RNA" else None,
            cdna_library_input if modality == "RNA" else None,
            library_prep_date,
            library_prep_set,
            library_name,
            rna_size if modality == "RNA" else atac_size,
            library_cycles,
            lib_quant,
            "Pass",
            f"SI-TT-{library_index}_i7" if modality == "RNA" else None,
            f"SI-TT-{library_index}_b(i5)" if modality == "RNA" else None,
            f"SI-NA-{library_index}" if modality == "ATAC" else None
        ]

        # ==========================================
        # UPDATED LOGIC FOR COLUMN V (cDNA counter)
        # Uses _next_amp_name to reconcile with sheet
        # so counters survive server restarts.
        # ==========================================
        if modality == "RNA":
            cdna_amp_date = self.convert_date(form_data.get('cdna_amp_date', ''))
            if not cdna_amp_date:
                cdna_amp_date = current_date  # Fallback if empty

            amp_date_key = f"amp_{cdna_amp_date}"
            amp_prefix = f"AP{experimenter_initials}{rna_suffix}"

            # --- Reconcile amp_counter with sheet data ---
            if amp_date_key not in state["amp_counter"]:
                # State doesn't know about this amp date — check the sheet
                next_name = self._next_amp_name(worksheet, amp_prefix, cdna_amp_date)
                # Parse the next_name to figure out what counter value it implies
                m = re.match(rf'^{re.escape(amp_prefix)}_{re.escape(cdna_amp_date)}_(\d+)_([A-H])$', next_name)
                if m:
                    batch = int(m.group(1))
                    letter = m.group(2)
                    # Convert batch+letter back to a reaction_count
                    letter_idx = ord(letter) - 65  # A=0, B=1, ...
                    state["amp_counter"][amp_date_key] = (batch - 1) * 8 + letter_idx
                else:
                    state["amp_counter"][amp_date_key] = 0

            reaction_count = state["amp_counter"][amp_date_key]

            letter = chr(65 + (reaction_count % 8))  # A through H
            batch_num_for_amp = (reaction_count // 8) + 1  # 1, then 2, then 3...

            row_data[21] = f"{amp_prefix}_{cdna_amp_date}_{batch_num_for_amp}_{letter}"

            state["amp_counter"][amp_date_key] += 1
        # ==========================================

        for col_num, value in enumerate(row_data, start=1):
            cell = worksheet.cell(row=current_row, column=col_num, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal='left')

            if (modality == "ATAC" and value is None) or (
                    modality == "RNA" and col_num == headers.index('ATAC_index') + 1):
                cell.fill = self.black_fill

        tissue_old_col = headers.index('tissue_name_old') + 1
        worksheet.cell(row=current_row, column=tissue_old_col).fill = self.black_fill

# favicon route (optional, for direct /favicon.ico requests)
@app.route('/favicon.ico')
def favicon():
    resp = make_response(
        send_from_directory(
            os.path.join(app.root_path, 'static'),
            'favicon.ico',
            mimetype='image/vnd.microsoft.icon'
        )
    )
    resp.headers['Cache-Control'] = 'public, max-age=31536000'
    return resp


data_logger = DataLogger()


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/update_counter', methods=['POST'])
def update_counter():
    try:
        request_data = request.json or {}
        # Accept user from JSON or query param
        user_first_name = (request_data.get('user_first_name') or request.args.get('user') or "").strip()
        if not user_first_name:
            return jsonify({'success': False, 'error': 'Missing user_first_name'}), 400

        # Accept counter as number or numeric string
        new_counter_raw = request_data.get('new_counter')
        try:
            new_counter = int(new_counter_raw)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': 'Invalid counter value'}), 400
        if new_counter < 0:
            return jsonify({'success': False, 'error': 'Invalid counter value'}), 400

        user_key = data_logger._safe_user_key(user_first_name)

        # If you’re persisting per-user state to GCS:
        # state = data_logger._load_user_state(user_key)
        # state['next_counter'] = new_counter
        # data_logger._save_user_state(user_key, state)

        # If you’re still using the local meta fallback:
        meta = data_logger._load_local_meta()
        states = meta.setdefault('user_states', {})
        state = states.setdefault(user_key, {"next_counter": None, "date_info": {}, "amp_counter": {}})
        state['next_counter'] = new_counter
        data_logger._save_local_meta(meta)

        return jsonify({'success': True, 'new_counter': new_counter})
    except Exception as e:
        # Surface server-side error to help diagnose
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/submit', methods=['POST'])
def submit_data():
    try:
        form_data = request.json
        required_fields = ['user_first_name', 'date', 'marmoset', 'slab', 'tile', 'hemisphere', 'tile_location', 'sort_method',
                           'rxn_number', 'sorter_initials']
        for field in required_fields:
            if not form_data.get(field):
                return jsonify({'success': False, 'error': f'Missing required field: {field}'})
        success = data_logger.process_form_data(form_data)
        return jsonify({'success': True, 'message': 'Data saved successfully!'}) if success else jsonify({'success': False, 'error': 'Failed to process data'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/download')
def download_excel():
    try:
        user_first_name = (request.args.get('user') or "").strip()
        if not user_first_name:
            return jsonify({'error': 'Missing user name in query parameter ?user='}), 400
        user_key = data_logger._safe_user_key(user_first_name)
        object_name = data_logger._load_pointer(user_key)
        if not object_name:
            object_name = data_logger._new_object_name(user_key)
            data_logger._save_pointer(user_key, object_name)
        data = data_logger._download_workbook_bytes(object_name)
        filename = os.path.basename(object_name)
        return send_file(
            io.BytesIO(data),
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500